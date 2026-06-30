"""Exportação do relatório "Vendas do Mês" em PDF (reportlab) e Excel (openpyxl)."""

from __future__ import annotations

import io

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
from reportlab.lib.units import mm
from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle

# Colunas: (rótulo, chave, alinhamento, formatador) — ordem do grid
_COLS = [
    ("Produto", "produto", "left", None),
    ("Vend.", "qtd_vendida", "center", None),
    ("Canc.", "qtd_cancelada", "center", None),
    ("Entr.", "qtd_entregue", "center", None),
    ("Custo", "custo", "right", "money"),
    ("Venda", "venda", "right", "money"),
    ("Lucro Bruto", "lucro_bruto", "right", "money"),
    ("% Lucro", "pct_lucro", "right", "pct"),
    ("Taxa (rat.)", "taxa_rateada", "right", "money"),
    ("Frete (rat.)", "frete_rateado", "right", "money"),
    ("LL Parcial", "ll_parcial", "right", "money"),
    ("% LL", "pct_ll_parcial", "right", "pct"),
]


def _money(v) -> str:
    try:
        s = f"{float(v):,.2f}"
    except (TypeError, ValueError):
        return "R$ 0,00"
    return "R$ " + s.replace(",", "X").replace(".", ",").replace("X", ".")


def _pct(v) -> str:
    try:
        return f"{float(v):.1f}%".replace(".", ",")
    except (TypeError, ValueError):
        return "0,0%"


def _produto(row: dict) -> str:
    titulo = (row.get("titulo") or "").strip()
    sku = (row.get("sku") or "").strip()
    if sku and titulo:
        return f"{titulo} ({sku})"
    return titulo or sku or "—"


def build_pdf(report: dict, account_label: str) -> bytes:
    buf = io.BytesIO()
    doc = SimpleDocTemplate(
        buf, pagesize=landscape(A4),
        leftMargin=10 * mm, rightMargin=10 * mm, topMargin=12 * mm, bottomMargin=12 * mm,
        title="Vendas do Mês",
    )
    styles = getSampleStyleSheet()
    h = ParagraphStyle("h", parent=styles["Title"], fontSize=14, spaceAfter=2)
    sub = ParagraphStyle("sub", parent=styles["Normal"], fontSize=9, textColor=colors.grey)
    cell = ParagraphStyle("cell", parent=styles["Normal"], fontSize=7, leading=8)
    cell_b = ParagraphStyle("cellb", parent=cell, fontName="Helvetica-Bold")

    story = [
        Paragraph("Vendas do Mês — por produto", h),
        Paragraph(f"{account_label} · período {report.get('period', '')}", sub),
        Spacer(1, 4 * mm),
    ]

    header = [Paragraph(f"<b>{label}</b>", cell_b) for label, *_ in _COLS]
    data = [header]
    for r in report.get("rows", []):
        line = []
        for _label, key, _align, fmt in _COLS:
            if key == "produto":
                line.append(Paragraph(_produto(r), cell))
            elif fmt == "money":
                line.append(Paragraph(_money(r.get(key)), cell))
            elif fmt == "pct":
                line.append(Paragraph(_pct(r.get(key)), cell))
            else:
                line.append(Paragraph(str(r.get(key, 0)), cell))
        data.append(line)

    t = report.get("totals", {})
    totals_line = [Paragraph("<b>TOTAL</b>", cell_b)]
    for _label, key, _align, fmt in _COLS[1:]:
        val = t.get(key, 0)
        txt = _money(val) if fmt == "money" else _pct(val) if fmt == "pct" else str(val)
        totals_line.append(Paragraph(f"<b>{txt}</b>", cell_b))
    data.append(totals_line)

    # larguras: Produto larga, demais estreitas
    widths = [95 * mm] + [(277 - 95) / (len(_COLS) - 1) * mm] * (len(_COLS) - 1)
    table = Table(data, colWidths=widths, repeatRows=1)
    aligns = [c[2].upper() for c in _COLS]
    style = [
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1f3b57")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("BACKGROUND", (0, -1), (-1, -1), colors.HexColor("#e8eef4")),
        ("GRID", (0, 0), (-1, -1), 0.25, colors.HexColor("#b9c4cf")),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("ROWBACKGROUNDS", (0, 1), (-1, -2), [colors.white, colors.HexColor("#f5f8fa")]),
        ("TOPPADDING", (0, 0), (-1, -1), 2),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 2),
    ]
    for i, a in enumerate(aligns):
        style.append(("ALIGN", (i, 0), (i, -1), a))
    table.setStyle(TableStyle(style))
    story.append(table)

    doc.build(story)
    return buf.getvalue()


def build_xlsx(report: dict, account_label: str) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Vendas do Mês"

    bold = Font(bold=True)
    head_fill = PatternFill("solid", fgColor="1F3B57")
    head_font = Font(bold=True, color="FFFFFF")
    tot_fill = PatternFill("solid", fgColor="E8EEF4")
    money_fmt = 'R$ #,##0.00'

    ws.append([f"Vendas do Mês — {account_label} · {report.get('period', '')}"])
    ws["A1"].font = bold
    ws.append([])

    head_row = 3
    ws.append([label for label, *_ in _COLS])
    for c in range(1, len(_COLS) + 1):
        cell = ws.cell(row=head_row, column=c)
        cell.fill = head_fill
        cell.font = head_font
        cell.alignment = Alignment(horizontal="center")

    def _write(values: list, *, total=False):
        ws.append(list(values))   # lista NÃO-vazia → avança a linha (append([]) não avançava)
        ri = ws.max_row
        for ci, (_label, key, align, fmt) in enumerate(_COLS, start=1):
            cell = ws.cell(row=ri, column=ci)
            cell.alignment = Alignment(horizontal=align)
            if fmt == "money":
                cell.number_format = money_fmt
            elif fmt == "pct":
                cell.number_format = '0.0"%"'
            if total:
                cell.font = bold
                cell.fill = tot_fill

    for r in report.get("rows", []):
        _write([
            _produto(r), r.get("qtd_vendida", 0), r.get("qtd_cancelada", 0), r.get("qtd_entregue", 0),
            r.get("custo", 0), r.get("venda", 0), r.get("lucro_bruto", 0), r.get("pct_lucro", 0),
            r.get("taxa_rateada", 0), r.get("frete_rateado", 0),
            r.get("ll_parcial", 0), r.get("pct_ll_parcial", 0),
        ])
    t = report.get("totals", {})
    _write([
        "TOTAL", t.get("qtd_vendida", 0), t.get("qtd_cancelada", 0), t.get("qtd_entregue", 0),
        t.get("custo", 0), t.get("venda", 0), t.get("lucro_bruto", 0), t.get("pct_lucro", 0),
        t.get("taxa_rateada", 0), t.get("frete_rateado", 0),
        t.get("ll_parcial", 0), t.get("pct_ll_parcial", 0),
    ], total=True)

    widths = [46, 9, 9, 9, 13, 13, 13, 9, 13, 13, 13, 9]
    for ci, w in enumerate(widths, start=1):
        ws.column_dimensions[ws.cell(row=head_row, column=ci).column_letter].width = w
    ws.freeze_panes = ws.cell(row=head_row + 1, column=1)

    out = io.BytesIO()
    wb.save(out)
    return out.getvalue()

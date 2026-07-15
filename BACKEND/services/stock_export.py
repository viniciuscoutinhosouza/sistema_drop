"""Exportação do Controle de Estoque em PDF (reportlab) e Excel (openpyxl).

Recebe a lista de itens já montada e escopada por `routers/stock._collect_stock_items`
(mesmos filtros/acesso da tela) e renderiza exatamente as colunas exibidas em
`StockControlView.vue`. Mesmo padrão de `integrations/eship/export.py` e
`services/sales_report_export.py` (dívida técnica: extrair um `table_export` genérico — ver LOG).
"""

from __future__ import annotations

import io

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
from reportlab.lib.units import mm
from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle

# (rótulo, chave, alinhamento, tipo) — ordem/colunas iguais às da tela de Controle de Estoque.
_COLS = [
    ("SKU", "sku", "left", "text"),
    ("Produto", "name", "left", "text"),
    ("EAN", "ean", "left", "text"),
    ("Tipo", "tipo_label", "left", "text"),
    ("Físico", "physical", "right", "num"),
    ("Reservado", "reserved", "right", "num"),
    ("Disponível", "available", "right", "num"),
    ("Ag.Retorno", "awaiting_return", "right", "num"),
    ("Ag.Validação", "pending_validation", "right", "num"),
    ("Inapto", "unfit", "right", "num"),
    ("FULL", "full_stock_total", "right", "num"),
]

_TITLE = "Controle de Estoque"


def _xlsx_safe(v):
    """Prefixa `'` em texto que começa com `= + - @` (anti formula-injection no Excel)."""
    if isinstance(v, str) and v[:1] in ("=", "+", "-", "@"):
        return "'" + v
    return v


def _cell_text(row: dict, key: str, tipo: str) -> str:
    v = row.get(key)
    if tipo == "num":
        return "0" if v is None else str(v)
    return "" if v is None else str(v)


def build_pdf(items: list[dict], subtitle: str) -> bytes:
    buf = io.BytesIO()
    doc = SimpleDocTemplate(
        buf, pagesize=landscape(A4),
        leftMargin=10 * mm, rightMargin=10 * mm, topMargin=12 * mm, bottomMargin=12 * mm,
        title=_TITLE,
    )
    styles = getSampleStyleSheet()
    h = ParagraphStyle("h", parent=styles["Title"], fontSize=14, spaceAfter=2)
    sub = ParagraphStyle("sub", parent=styles["Normal"], fontSize=9, textColor=colors.grey)
    cell = ParagraphStyle("cell", parent=styles["Normal"], fontSize=7, leading=8)
    cell_b = ParagraphStyle("cellb", parent=cell, fontName="Helvetica-Bold")

    story = [Paragraph(_TITLE, h), Paragraph(subtitle, sub), Spacer(1, 4 * mm)]

    header = [Paragraph(f"<b>{label}</b>", cell_b) for label, *_ in _COLS]
    table_data = [header]
    for r in items:
        table_data.append([
            Paragraph(_cell_text(r, key, tipo).replace("&", "&amp;").replace("<", "&lt;"), cell)
            for _label, key, _align, tipo in _COLS
        ])
    if len(table_data) == 1:  # só header → nenhum item
        table_data.append([Paragraph("Nenhum produto.", cell)] + [Paragraph("", cell)] * (len(_COLS) - 1))

    # A4 paisagem útil ≈ 277mm: produto largo, colunas numéricas estreitas.
    widths = [28, 62, 26, 34, 16, 20, 18, 20, 22, 16, 15]
    table = Table(table_data, colWidths=[w * mm for w in widths], repeatRows=1)
    style = [
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1f3b57")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("GRID", (0, 0), (-1, -1), 0.25, colors.HexColor("#b9c4cf")),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#f5f8fa")]),
        ("TOPPADDING", (0, 0), (-1, -1), 2),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 2),
    ]
    for i, (_label, _key, align, _tipo) in enumerate(_COLS):
        style.append(("ALIGN", (i, 0), (i, -1), align.upper()))
    table.setStyle(TableStyle(style))
    story.append(table)

    doc.build(story)
    return buf.getvalue()


def build_xlsx(items: list[dict], subtitle: str) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Estoque"

    bold = Font(bold=True)
    head_fill = PatternFill("solid", fgColor="1F3B57")
    head_font = Font(bold=True, color="FFFFFF")

    ws.append([f"{_TITLE} — {subtitle}"])
    ws["A1"].font = bold
    ws.append([])

    head_row = 3
    ws.append([label for label, *_ in _COLS])
    for c in range(1, len(_COLS) + 1):
        hc = ws.cell(row=head_row, column=c)
        hc.fill = head_fill
        hc.font = head_font
        hc.alignment = Alignment(horizontal="center")

    for r in items:
        values = []
        for _label, key, _align, tipo in _COLS:
            if tipo == "num":
                values.append(r.get(key) if r.get(key) is not None else 0)
            else:
                values.append(_xlsx_safe(r.get(key) or ""))
        ws.append(values)
        ri = ws.max_row
        for ci, (_label, _key, align, _tipo) in enumerate(_COLS, start=1):
            ws.cell(row=ri, column=ci).alignment = Alignment(horizontal=align)

    widths = [18, 50, 16, 26, 9, 11, 11, 12, 13, 9, 8]
    for ci, w in enumerate(widths, start=1):
        ws.column_dimensions[ws.cell(row=head_row, column=ci).column_letter].width = w
    ws.freeze_panes = ws.cell(row=head_row + 1, column=1)

    out = io.BytesIO()
    wb.save(out)
    return out.getvalue()

"""Exportação da lista de produtos do eShip (por empresa) em PDF (reportlab) e Excel (openpyxl).

Recebe o dict de `service.list_all_eship_products` (já escopado à CMIG e com estoque real) e
renderiza as colunas exibidas na tela: código, cód. barras, descrição, status, Full, físico,
disponível e reservado. Mesmo padrão de `services/sales_report_export.py`.
"""

from __future__ import annotations

import io

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4
from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
from reportlab.lib.units import mm
from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle

# (rótulo, chave, alinhamento, tipo) — ordem/colunas iguais às da tela
_COLS = [
    ("Código", "codigo", "left", "text"),
    ("Cód. barras", "codigo_barras", "left", "text"),
    ("Descrição", "descricao", "left", "text"),
    ("Status", "status", "left", "text"),
    ("Full", "is_full", "center", "full"),
    ("Físico", "total_fisico", "right", "num"),
    ("Disp.", "total_disponivel", "right", "num"),
    ("Reserv.", "total_reservado", "right", "num"),
]


def _xlsx_safe(v):
    """Prefixa `'` em texto que começa com `= + - @` (anti formula-injection no Excel)."""
    if isinstance(v, str) and v[:1] in ("=", "+", "-", "@"):
        return "'" + v
    return v


def _cell_text(row: dict, key: str, tipo: str) -> str:
    v = row.get(key)
    if tipo == "full":
        return "Sim" if v else "—"
    if tipo == "num":
        return "0" if v is None else str(v)
    return "" if v is None else str(v)


def _subtitle(data: dict, cmig_label: str) -> str:
    total = data.get("total", 0)
    cat = data.get("total_catalogo")
    extra = f" · catálogo WMS: {cat}" if cat is not None else ""
    return f"{cmig_label} · {total} produto(s){extra}"


def build_pdf(data: dict, cmig_label: str) -> bytes:
    buf = io.BytesIO()
    doc = SimpleDocTemplate(
        buf, pagesize=A4,
        leftMargin=10 * mm, rightMargin=10 * mm, topMargin=12 * mm, bottomMargin=12 * mm,
        title="Produtos no eShip",
    )
    styles = getSampleStyleSheet()
    h = ParagraphStyle("h", parent=styles["Title"], fontSize=14, spaceAfter=2)
    sub = ParagraphStyle("sub", parent=styles["Normal"], fontSize=9, textColor=colors.grey)
    cell = ParagraphStyle("cell", parent=styles["Normal"], fontSize=7, leading=8)
    cell_b = ParagraphStyle("cellb", parent=cell, fontName="Helvetica-Bold")

    story = [
        Paragraph("Produtos no eShip", h),
        Paragraph(_subtitle(data, cmig_label), sub),
        Spacer(1, 4 * mm),
    ]

    header = [Paragraph(f"<b>{label}</b>", cell_b) for label, *_ in _COLS]
    table_data = [header]
    for r in data.get("produtos", []):
        table_data.append([
            Paragraph(_cell_text(r, key, tipo).replace("&", "&amp;").replace("<", "&lt;"), cell)
            for _label, key, _align, tipo in _COLS
        ])
    if len(table_data) == 1:  # só header → nenhum produto
        table_data.append([Paragraph("Nenhum produto.", cell)] + [Paragraph("", cell)] * (len(_COLS) - 1))

    # larguras (A4 retrato útil ≈ 190mm): descrição larga
    widths = [22 * mm, 22 * mm, 72 * mm, 20 * mm, 11 * mm, 15 * mm, 14 * mm, 14 * mm]
    table = Table(table_data, colWidths=widths, repeatRows=1)
    style = [
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1f3b57")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("GRID", (0, 0), (-1, -1), 0.25, colors.HexColor("#b9c4cf")),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#f5f8fa")]),
        ("TOPPADDING", (0, 0), (-1, -1), 2),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 2),
    ]
    for i, (_label, _key, align, _tipo) in enumerate(_COLS):
        style.append(("ALIGN", (i, 0), (i, -1), align.upper()))
    table.setStyle(TableStyle(style))
    story.append(table)

    doc.build(story)
    return buf.getvalue()


def build_xlsx(data: dict, cmig_label: str) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Produtos eShip"

    bold = Font(bold=True)
    head_fill = PatternFill("solid", fgColor="1F3B57")
    head_font = Font(bold=True, color="FFFFFF")

    ws.append([f"Produtos no eShip — {_subtitle(data, cmig_label)}"])
    ws["A1"].font = bold
    ws.append([])

    head_row = 3
    ws.append([label for label, *_ in _COLS])
    for c in range(1, len(_COLS) + 1):
        hc = ws.cell(row=head_row, column=c)
        hc.fill = head_fill
        hc.font = head_font
        hc.alignment = Alignment(horizontal="center")

    for r in data.get("produtos", []):
        values = []
        for _label, key, _align, tipo in _COLS:
            if tipo == "full":
                values.append("Sim" if r.get(key) else "")
            elif tipo == "num":
                values.append(r.get(key) if r.get(key) is not None else 0)
            else:
                values.append(_xlsx_safe(r.get(key) or ""))
        ws.append(values)
        ri = ws.max_row
        for ci, (_label, _key, align, _tipo) in enumerate(_COLS, start=1):
            ws.cell(row=ri, column=ci).alignment = Alignment(horizontal=align)

    widths = [16, 16, 60, 14, 6, 10, 10, 10]
    for ci, w in enumerate(widths, start=1):
        ws.column_dimensions[ws.cell(row=head_row, column=ci).column_letter].width = w
    ws.freeze_panes = ws.cell(row=head_row + 1, column=1)

    out = io.BytesIO()
    wb.save(out)
    return out.getvalue()

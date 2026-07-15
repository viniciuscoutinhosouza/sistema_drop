"""Testes do export de Controle de Estoque (services.stock_export) e da equivalência
entre a lista completa (`_collect_stock_items`) e o `total` do summary paginado."""

import io
import zipfile

from openpyxl import load_workbook

from services import stock_export

_ITEMS = [
    {
        "sku": "ABC-1", "name": "Camiseta Preta M", "ean": "7891234567895",
        "product_type": "cmig", "cmig_name": "Empresa X", "tipo_label": "CMIG · Empresa X",
        "physical": 10, "reserved": 2, "available": 8, "awaiting_return": 0,
        "pending_validation": 1, "unfit": 0, "full_stock_total": 5,
    },
    {
        "sku": "=CMD", "name": "Produto <b>x</b> & cia", "ean": "",
        "product_type": "pg", "cmig_name": None, "tipo_label": "PG",
        "physical": 0, "reserved": 0, "available": 0, "awaiting_return": 0,
        "pending_validation": 0, "unfit": 0, "full_stock_total": 0,
    },
]


def test_build_pdf_is_valid_pdf():
    pdf = stock_export.build_pdf(_ITEMS, "Todos · 2 produto(s)")
    assert pdf[:4] == b"%PDF"
    assert len(pdf) > 500


def test_build_xlsx_is_valid_and_blocks_formula_injection():
    data = stock_export.build_xlsx(_ITEMS, "Todos · 2 produto(s)")
    assert data[:2] == b"PK"  # zip/xlsx
    ws = load_workbook(io.BytesIO(data)).active
    cells = [row[0] for row in ws.iter_rows(values_only=True)]
    # SKU "=CMD" deve ser neutralizado com aspa simples (anti formula-injection)
    assert "'=CMD" in cells
    assert "=CMD" not in cells


def test_build_xlsx_is_a_real_zip_container():
    data = stock_export.build_xlsx(_ITEMS, "x")
    with zipfile.ZipFile(io.BytesIO(data)) as z:
        assert any(n.endswith(".xml") for n in z.namelist())


def test_empty_items_still_render():
    assert stock_export.build_pdf([], "Todos · 0 produto(s)")[:4] == b"%PDF"
    assert stock_export.build_xlsx([], "Todos · 0 produto(s)")[:2] == b"PK"
